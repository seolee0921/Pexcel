import cv2
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Color
from openpyxl.utils.cell import get_column_letter

wb = Workbook()
ws = wb.active

image = cv2.imread("teemo minimalistic.jpg")

print(image.shape)
print(image[100][100])

def _from_rgb(rgb):
    return "%02x%02x%02x" % rgb

for x in range(1, int(image.shape[1] / 9)):
    cell = get_column_letter(x)
    ws.column_dimensions[cell].width = 3

for y in range(1, int(image.shape[0] / 9)):
    for x in range(1, int(image.shape[1] / 9)):
        (B, G, R) = image[y * 9][x * 9]
        ws.cell(row=y, column=x).fill = PatternFill(fgColor=_from_rgb((R, G, B)), fill_type='solid')


print("SUCESS")
wb.save("print-picture.xlsx")