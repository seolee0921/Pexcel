import cv2
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter

wb = Workbook()
ws = wb.active

pixel = 8

def _from_rgb(rgb):
    return "%02x%02x%02x" % rgb

CAM_ID = 0

def capture(camid=CAM_ID):
    cam = cv2.VideoCapture(camid, cv2.CAP_DSHOW)
    cam.set(3, 1920)
    cam.set(4, 1080)
    if cam.isOpened() == False:
        print('cant open the cam (%d)' % camid)
        return None

    ret, frame = cam.read()
    if frame is None:
        print('frame is not exist')
        return None

    cv2.imwrite('save_video.png', frame, params=[cv2.IMWRITE_PNG_COMPRESSION, 0])
    cam.release()


capture()

image = cv2.imread("save_video.png")

for y in range(1, int(image.shape[0] / pixel)):
    for x in range(1, int(image.shape[1] / pixel)):
        cell = get_column_letter(x)
        ws.column_dimensions[cell].width = 3
        (B, G, R) = image[y * pixel][x * pixel]
        ws.cell(row=y, column=x).fill = PatternFill(fgColor=_from_rgb((R, G, B)), fill_type='solid')

wb.save("print-picture.xlsx")

open("print-picture.xlsx")
print("SUCCESS")
